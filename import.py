import os
import pandas as pd
from PIL import Image
from PIL.ExifTags import TAGS, GPSTAGS
from io import BytesIO
import base64
from tkinter import Tk, filedialog
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage

# Funktion zum Extrahieren von EXIF-Daten (GPS, Datum, Name)
def get_exif_data(image_path):
    image = Image.open(image_path)
    exif_data = image._getexif()
    data = {
        'Filename': os.path.basename(image_path),
        'Date': None,
        'Time': None,
        'Latitude': None,
        'Longitude': None,
        'Image': None
    }
    if exif_data:
        for tag, value in exif_data.items():
            tag_name = TAGS.get(tag, tag)
            if tag_name == 'DateTime':
                if value:
                    date_time = value.split()
                    if len(date_time) == 2:
                        data['Date'] = date_time[0]
                        data['Time'] = date_time[1]
            if tag_name == 'GPSInfo':
                gps_info = {}
                for gps_tag, gps_value in value.items():
                    gps_tag_name = GPSTAGS.get(gps_tag, gps_tag)
                    gps_info[gps_tag_name] = gps_value
                if 'GPSLatitude' in gps_info and 'GPSLongitude' in gps_info:
                    lat = convert_to_degrees(gps_info['GPSLatitude'])
                    lon = convert_to_degrees(gps_info['GPSLongitude'])
                    if gps_info.get('GPSLatitudeRef') == 'S':
                        lat = -lat
                    if gps_info.get('GPSLongitudeRef') == 'W':
                        lon = -lon
                    data['Latitude'] = f'{lat:.6f}°'
                    data['Longitude'] = f'{lon:.6f}°'
    # Bild in Base64 umwandeln
    image.thumbnail((100, 100))
    buffer = BytesIO()
    image.save(buffer, format="JPEG")
    encoded_image = base64.b64encode(buffer.getvalue()).decode('utf-8')
    data['Image'] = f'data:image/jpeg;base64,{encoded_image}'
    return data

def convert_to_degrees(value):
    d, m, s = value
    return float(d) + (float(m) / 60.0) + (float(s) / 3600.0)

# Ordner auswählen
def select_folder():
    root = Tk()
    root.withdraw()
    folder_selected = filedialog.askdirectory(title="Ordner auswählen")
    return folder_selected

# Zielordner auswählen
def select_output_folder():
    root = Tk()
    root.withdraw()
    folder_selected = filedialog.askdirectory(title="Speicherort auswählen")
    return folder_selected

# Hauptfunktion
def main():
    folder = select_folder()
    if not folder:
        print("Kein Ordner ausgewählt.")
        return

    output_folder = select_output_folder()
    if not output_folder:
        print("Kein Speicherort ausgewählt.")
        return

    data_list = []
    image_files = []
    for filename in os.listdir(folder):
        if filename.lower().endswith(('jpg', 'jpeg', 'png')):
            file_path = os.path.join(folder, filename)
            try:
                data = get_exif_data(file_path)
                data_list.append(data)
                image_files.append(file_path)  # Speichern der Bilddateipfade für späteres Hinzufügen
            except Exception as e:
                print(f"Fehler bei Datei {filename}: {e}")

    df = pd.DataFrame(data_list)

    # Excel-Datei erstellen
    wb = Workbook()
    ws = wb.active
    ws.title = "Bilder"

    # Kopfzeilen
    ws.append(['Vorschau', 'Name', 'Datum', 'Uhrzeit', 'Latitude', 'Longitude'])

    # Bilder und Daten hinzufügen
    for index, row in df.iterrows():
        # Bildvorschau einfügen
        img_path = image_files[index]
        img = Image.open(img_path)
        img.thumbnail((100, 100))  # Thumbnail erstellen

        # Bild in Excel einfügen
        img_bytes = BytesIO()
        img.save(img_bytes, format="JPEG")
        img_bytes.seek(0)
        excel_img = ExcelImage(img_bytes)
        
        # Bild in Excel hinzufügen
        img_cell = f'A{index + 2}'
        ws.add_image(excel_img, img_cell)
        
        # Daten hinzufügen
        ws[f'B{index + 2}'] = row['Filename']
        ws[f'C{index + 2}'] = row['Date']
        ws[f'D{index + 2}'] = row['Time']
        ws[f'E{index + 2}'] = row['Latitude']
        ws[f'F{index + 2}'] = row['Longitude']

    # Speicherort für die Excel-Datei
    output_path = os.path.join(output_folder, 'Bilder_Liste.xlsx')
    wb.save(output_path)

    print(f"Datei erstellt: {output_path}")

if __name__ == "__main__":
    main()
